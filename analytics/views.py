from datetime import datetime

from django.views.generic import TemplateView
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse

# libreria para el manejo de archivos xls
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Fill

# Models
from users.models import User
from home.models import Question
class ReporteUsuarios(TemplateView):
    """
    Vista basada en clases que retorna un archivo de reporte de formato xls (Excel)
        HttpResponse de content_type='application/ms-excel'
    Para obtener archivo, debe ser llamado desde petición get
    """
    def get(self, request, *args, **kwargs):
        """
        Funcionamiento:
            Para la generación de los campos, unicamente se deben cargar los modelos
            y sus campos respectivos, así como sus inner joins que se requieren para 
            mostrarse en el excel.
        """
        # Se cargan los modelos para insertar la información
        users = User.objects.all()
        # Se el objeto xls
        workbook = Workbook()
        # Contador para iniciar desde la fila 4 a colocar la información de los modelos   
        count = 3
        # Activa la hoja de excel 1 para trabajar
        ws1 = workbook.active
        ws2 = workbook.create_sheet()
        # Inserta contenido a celdas
        ws1.title = "Reporte de usuarios"
        ws2.title = "Reporte de preguntas"
        ws1['A1'] = 'Reporte general de usuarios Albéitar'
        #ws1.column_dimensions['D3'].width = 150
        # Aplica estilos a las celdas
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('A1:D1')
        
        # Se definen los encabezados de las columnas
        ws1['A2'] = 'ID'
        ws1['B2'] = 'Usuario'
        ws1['C2'] = 'Rol'
        ws1['D2'] = 'Especialidad'
        ws1.column_dimensions["D"].width = 25
        c = ws1['D2']
        c.font = Font(size=12)
        for user in users:
            # Itera celdas y añade contenido
            ws1.cell(row=count, column=1, value=user.id)
            ws1.cell(row=count, column=2, value=user.username)
            ws1.cell(row=count, column=3, value=user.get_rol_display())
            if not user.speciality:
                ws1.cell(row=count, column=4, value='S/E')
            else:
                ws1.cell(row=count, column=4, value=user.speciality)
            count += 1

        ws2['A1'] = 'Reporte general de preguntas'
        # Aplica estilos a las celdas
        ws2['A1'].alignment = Alignment(horizontal='center')
        ws2.merge_cells('A1:E1')
        count = 3
        # Se definen los encabezados de las columnas
        ws2['A2'] = 'ID'
        ws2['B2'] = 'Titulo'
        ws2['C2'] = 'Especie'
        ws2['D2'] = 'Alumno'
        ws2['E2'] = 'Estado'
        c = ws2['A1']
        c.font = Font(size=12, bold = True)
        c = ws1['A1']
        c.font = Font(size=12, bold=True)
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 25
        ws2.column_dimensions["D"].width = 18
        ws2.column_dimensions["E"].width = 18
        questions = Question.objects.all()
        for question in questions:
            # Itera celdas y añade contenido
            ws2.cell(row=count, column=1, value=question.id)
            ws2.cell(row=count, column=2, value=question.title)
            ws2.cell(row=count, column=3, value=question.get_specie_display())
            ws2.cell(row=count, column=4, value=question.user_question.username)
            ws2.cell(row=count, column=5, value=question.get_status_display())
            count += 1


        file_name = 'Reporte_General_De_Usuarios_{0}.xlsx'.format(datetime.now().strftime("%I-%M%p_%d-%m-%Y"))
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename={0}'.format(file_name)
        response['Content-Disposition'] = content
        workbook.save(response)
        if request.user.rol == 'AD':
            return response
        else:
            return redirect('home:inicio')
